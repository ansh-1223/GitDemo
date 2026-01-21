import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

#edit the excel with updated value
def updateExcel(file, searchTerm, colName, new_value):
    book=openpyxl.load_workbook(file)
    sheet=book.active
    for i in range(1,sheet.max_row+1): #for rows
        if sheet.cell(row=i,column=2).value ==searchTerm:
            for j in range(3,sheet.max_column+1):
                if sheet.cell(row=1,column=j).value ==colName:
                    sheet.cell(row=i, column=j).value = new_value
    book.save(file_path)
    time.sleep(5)


fruit_name="Apple"
new_value=999
file_path="C:\\Users\\AnshumanSingh\\Downloads\\download.xlsx"
driver=webdriver.Chrome()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/")
driver.find_element(By.ID,"downloadButton").click()

#edit the excel with updated value
updateExcel(file_path,"Apple","price",new_value)
file_input=driver.find_element(By.CSS_SELECTOR,"input[type='file']")
file_input.send_keys(file_path) #send keys is used to send the file as input

wait=WebDriverWait(driver,10)
toast_locator=(By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text) # * is used to unpack the variable values as it is not a string

#data column id is common for the column
price_column=driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
#parent:: is used to go up the hierarchy,here we first go to the name of fruit then go up to parent div and then traverse to the price column
actual_price=driver.find_element(By.XPATH,"//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+price_column+"-undefined']")
print(actual_price.text)
assert int(actual_price.text) == new_value
time.sleep(100)

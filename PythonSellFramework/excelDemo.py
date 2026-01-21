import openpyxl

book=openpyxl.load_workbook("C:\\Users\\AnshumanSingh\\Desktop\\PythonDemo2.xlsx")
Dict={}
sheet=book.active
cell=sheet.cell(row=1,column=2)
print(cell.value)
sheet.cell(row=2,column=2).value="Anshu"
print(sheet.cell(row=2,column=2).value)
book.save("C:\\Users\\AnshumanSingh\\Desktop\\PythonDemo2.xlsx")
print(sheet.max_row)
print(sheet.max_column)
print(sheet['A5'].value)

for i in range(1,sheet.max_row+1): #for rows
    if sheet.cell(row=i,column=1).value =="testcase2":
        for j in range(2,sheet.max_column+1):
            #print(sheet.cell(row=i,column=j).value)
            Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value
print(Dict)